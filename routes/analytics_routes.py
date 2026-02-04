from __future__ import annotations

import csv
import io
import json
import os
import unicodedata
from datetime import date, datetime, timedelta

from sqlalchemy import func

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response, current_app, send_file
from flask_login import login_required, current_user

from models.extensions import db
from models.entrada_model import Entrada
from models.projection_scenario_model import ProjectionScenario
from models.recurrence_model import Recurrence, RecurrenceExecution
from services.projection_engine import compute_projection
from services.plans import PLANS, is_valid_plan
from services.feature_gate import require_feature
from services.permissions import require_api_access, json_error, require_verified_email
from services.security import safe_redirect_path
from services.checkout_store import (
    create_order,
    set_order_billing_id,
    get_order_by_token,
    mark_order_paid_by_billing_id,
    mark_order_paid_by_token,
    list_orders_by_user,
)
from services.abacatepay import create_plan_billing, get_billing_status, AbacatePayError, payment_warning_message
from services.date_utils import last_day_of_month
from services.subscription import apply_paid_order
from services.reports_pdf import render_reports_pdf
from services.document_validation import (
    normalize_cpf,
    normalize_phone,
    validate_cpf,
    validate_phone,
)


analytics_bp = Blueprint("analytics", __name__)

CATEGORIAS = {
    "salario": "Salário",
    "extras": "Extras",
    "moradia": "Moradia",
    "mercado": "Mercado",
    "transporte": "Transporte",
    "servicos": "Serviços",
    "outros": "Outros",
}

STATUS_PADROES = {"pago", "em_andamento", "nao_pago"}

MODE_LABELS = {
    "cash": "Caixa",
    "accrual": "Competência",
}

DEFAULT_REPORT_SECTIONS = {"summary", "dre", "flow", "categories", "recurring", "pending"}
FLOW_LIMIT_RESUMIDO = 45
FLOW_LIMIT_DETALHADO = None

MESES = [
    "Janeiro",
    "Fevereiro",
    "Março",
    "Abril",
    "Maio",
    "Junho",
    "Julho",
    "Agosto",
    "Setembro",
    "Outubro",
    "Novembro",
    "Dezembro",
]


def _normalize_categoria(value: str | None) -> str:
    categoria = (value or "").strip().lower()
    if categoria not in CATEGORIAS:
        return "outros"
    return categoria


def _parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _summary_for_period(start: date, end: date) -> dict:
    entries = (
        Entrada.query
        .filter(
            Entrada.user_id == current_user.id,
            Entrada.data >= start,
            Entrada.data <= end,
        )
        .all()
    )

    receitas_total = sum(float(e.valor) for e in entries if e.tipo == "receita")
    despesas_total = sum(float(e.valor) for e in entries if e.tipo == "despesa")
    receitas_count = sum(1 for e in entries if e.tipo == "receita")
    despesas_count = sum(1 for e in entries if e.tipo == "despesa")

    saldo_projetado = receitas_total - despesas_total

    categoria_totais = {key: 0.0 for key in CATEGORIAS}
    for e in entries:
        if e.tipo != "despesa":
            continue
        cat = _normalize_categoria(getattr(e, "categoria", None))
        categoria_totais[cat] += float(e.valor)

    categorias = []
    for key, total in categoria_totais.items():
        if total <= 0:
            continue
        percent = (total / despesas_total * 100) if despesas_total else 0.0
        categorias.append(
            {
                "key": key,
                "label": CATEGORIAS.get(key, "Outros"),
                "total": round(total, 2),
                "percent": round(percent, 1),
            }
        )
    categorias.sort(key=lambda item: item["total"], reverse=True)

    top_receitas = sorted(
        [e for e in entries if e.tipo == "receita"],
        key=lambda item: float(item.valor),
        reverse=True,
    )[:10]
    top_despesas = sorted(
        [e for e in entries if e.tipo == "despesa"],
        key=lambda item: float(item.valor),
        reverse=True,
    )[:10]

    def _entry_payload(entry: Entrada) -> dict:
        return {
            "id": entry.id,
            "date": entry.data.isoformat() if entry.data else None,
            "type": entry.tipo,
            "description": entry.descricao,
            "category": _normalize_categoria(getattr(entry, "categoria", None)),
            "value": round(float(entry.valor), 2),
            "status": entry.status,
        }

    return {
        "period": {"start": start.isoformat(), "end": end.isoformat()},
        "summary": {
            "receitas": round(receitas_total, 2),
            "despesas": round(despesas_total, 2),
            "saldo_projetado": round(saldo_projetado, 2),
            "entradas": len(entries),
            "receitas_count": receitas_count,
            "despesas_count": despesas_count,
        },
        "categories": categorias,
        "top_entries": [_entry_payload(item) for item in top_receitas]
        + [_entry_payload(item) for item in top_despesas],
    }


def _normalize_text(value: str | None) -> str:
    text = (value or "").strip().lower()
    if not text:
        return ""
    text = "".join(
        char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn"
    )
    return text


def _normalize_method(value: str | None) -> str:
    text = _normalize_text(value)
    if "cart" in text:
        if "deb" in text:
            return "debito"
        if "cred" in text:
            return "credito"
        return "cartao"
    return text


def _method_matches(value: str | None, allowed: set[str]) -> bool:
    if not allowed:
        return True
    norm = _normalize_method(value)
    if not norm:
        return False
    if norm in allowed:
        return True
    # fallback: cartao conta como credito ou debito se filtrado
    if norm == "cartao" and ("credito" in allowed or "debito" in allowed):
        return True
    if norm in {"credito", "debito"} and "cartao" in allowed:
        return True
    return False


METHOD_LABELS = {
    "pix": "PIX",
    "credito": "Crédito",
    "debito": "Débito",
    "dinheiro": "Dinheiro",
    "boleto": "Boleto",
    "cartao": "Cartão",
}

STATUS_LABELS = {
    "pago": "Pago",
    "recebido": "Recebido",
    "em_andamento": "Em andamento",
    "nao_pago": "Nao pago",
}


def _method_label(value: str | None) -> str:
    if not value:
        return ""
    norm = _normalize_method(value)
    if norm in METHOD_LABELS:
        return METHOD_LABELS[norm]
    return value.strip().title()


def _status_label(value: str | None) -> str:
    if not value:
        return ""
    norm = value.strip().lower()
    if norm in STATUS_LABELS:
        return STATUS_LABELS[norm]
    return value.strip().title()


def _parse_list_param(value: str | None) -> set[str]:
    if not value:
        return set()
    items = set()
    for part in value.split(","):
        item = _normalize_text(part)
        if item:
            items.add(item)
    return items


def _format_period_label(start: date, end: date) -> str:
    return f"{start.strftime('%d/%m/%Y')} - {end.strftime('%d/%m/%Y')}"


def _resolve_report_period(period: str | None, start_str: str | None, end_str: str | None) -> tuple[date, date]:
    today = date.today()
    period = (period or "month").strip().lower()

    if period == "30":
        start = today - timedelta(days=29)
        end = today
    elif period == "quarter":
        quarter = (today.month - 1) // 3
        start_month = quarter * 3 + 1
        start = date(today.year, start_month, 1)
        end = last_day_of_month(date(today.year, start_month + 2, 1))
    elif period == "year":
        start = date(today.year, 1, 1)
        end = date(today.year, 12, 31)
    elif period == "custom":
        start = _parse_iso_date(start_str) or date(today.year, today.month, 1)
        end = _parse_iso_date(end_str) or last_day_of_month(start)
    else:
        start = date(today.year, today.month, 1)
        end = last_day_of_month(start)

    if end < start:
        start, end = end, start
    return start, end


def _entry_event_date(entry: Entrada, mode: str) -> date | None:
    if mode == "cash":
        if entry.tipo == "receita":
            return entry.received_at
        return entry.paid_at
    return entry.data


def _entry_amount(entry: Entrada) -> float:
    value = float(entry.valor or 0.0)
    return value if entry.tipo == "receita" else -value


def _filter_entries_for_period(
    entries: list[Entrada],
    start: date,
    end: date,
    mode: str,
    type_filter: str,
    status_filter: str,
    categories: set[str],
    methods: set[str],
) -> list[tuple[Entrada, date]]:
    items: list[tuple[Entrada, date]] = []
    for entry in entries:
        if type_filter == "income" and entry.tipo != "receita":
            continue
        if type_filter == "expense" and entry.tipo != "despesa":
            continue

        event_date = _entry_event_date(entry, mode)
        if not event_date:
            continue

        if event_date < start or event_date > end:
            continue

        if categories:
            cat = _normalize_categoria(getattr(entry, "categoria", None))
            if cat not in categories:
                continue

        if not _method_matches(getattr(entry, "metodo", None), methods):
            continue

        status = (entry.status or "").strip().lower()
        if status_filter == "paid":
            if status not in {"pago", "recebido"}:
                continue
        elif status_filter == "pending":
            if status in {"pago", "recebido"}:
                continue

        items.append((entry, event_date))
    return items


def _history_from_orders(orders):
    items = []
    for order in orders or []:
        plan = PLANS.get(order.plan, PLANS["basic"])
        when = order.paid_at or order.created_at
        date = when.strftime("%d/%m/%Y") if when else "-"
        items.append(
            {
                "date": date,
                "plan": plan["name"],
                "status": (order.status or "").upper() or "PENDING",
                "amount": plan.get("price_month"),
            }
        )
    return items


@analytics_bp.get("/app/upgrade")
@login_required
def upgrade():
    """Página de planos (upgrade/downgrade) via checkout."""
    return render_template("upgrade.html")


@analytics_bp.get("/app/upgrade/checkout")
@login_required
@require_verified_email("Confirme seu email para liberar o pagamento.")
def upgrade_checkout_page():
    plan = (request.args.get("plan") or "basic").strip().lower()
    if not is_valid_plan(plan):
        plan = "basic"
    if plan == current_user.plan:
        flash("Você já está nesse plano.", "info")
        return redirect(url_for("analytics.upgrade"))
    return render_template("upgrade_checkout.html", selected_plan=plan)


@analytics_bp.post("/app/upgrade/checkout")
@login_required
@require_verified_email("Confirme seu email para liberar o pagamento.")
def upgrade_checkout_start():
    plan = (request.form.get("plan") or "basic").strip().lower()
    if not is_valid_plan(plan):
        plan = "basic"
    if plan == current_user.plan:
        flash("Você já está nesse plano.", "info")
        return redirect(url_for("analytics.upgrade"))

    # Cria pedido local vinculado ao usuário logado
    order = create_order(plan, user_id=current_user.id)

    # Cria cobrança no provedor (PIX)
    try:
        # Para iniciar o pagamento, precisamos de dados pessoais válidos.
        full_name = (current_user.full_name or "").strip()
        tax_id = normalize_cpf(getattr(current_user, "tax_id", None))
        phone = normalize_phone(getattr(current_user, "cellphone", None))
        if not full_name:
            flash("Informe seu nome completo para liberar o pagamento.", "info")
            return redirect(url_for("account_page", section="profile"))
        if not validate_cpf(tax_id) or not validate_phone(phone):
            flash("Informe um CPF e telefone válidos para liberar o pagamento.", "info")
            return redirect(url_for("account_page", section="profile"))

        # Mantemos o mesmo padrão do checkout público:
        # - completionUrl: página que o cliente verá após pagar
        # - returnUrl: fallback/voltar
        completion_url = url_for("analytics.upgrade_return", token=order.token, _external=True)
        return_url = url_for("analytics.upgrade", _external=True)

        customer = {
            "name": full_name,
            "email": current_user.email,
            "cellphone": phone,
            "taxId": tax_id,
        }

        billing = create_plan_billing(
            plan=plan,
            external_id=order.token,
            return_url=return_url,
            completion_url=completion_url,
            customer=customer,
        )
    except AbacatePayError as e:
        flash(str(e), "error")
        return redirect(url_for("analytics.upgrade_checkout_page", plan=plan))
    except Exception:
        flash("Não foi possível iniciar o checkout agora. Tente novamente.", "error")
        return redirect(url_for("analytics.upgrade_checkout_page", plan=plan))

    set_order_billing_id(order.token, billing["billing_id"])
    return redirect(billing["url"])


@analytics_bp.get("/app/upgrade/return")
@login_required
@require_verified_email("Confirme seu email para liberar o pagamento.")
def upgrade_return():
    """Página de retorno pós-pagamento (cliente redirecionado pelo provedor)."""
    token = (request.args.get("token") or "").strip()
    order = get_order_by_token(token)
    if not order:
        flash("Nao encontramos esse checkout.", "error")
        return redirect(url_for("analytics.upgrade"))
    if not order.user_id or int(order.user_id) != int(current_user.id):
        flash("Checkout nao encontrado para esta conta.", "error")
        return redirect(url_for("analytics.upgrade"))

    # Se o webhook já marcou como pago, aplica o plano aqui.
    if order.status == "PAID":
        if apply_paid_order(current_user, order):
            db.session.commit()
        flash(f"Pagamento confirmado. Seu plano foi atualizado para {PLANS[order.plan]['name']}.", "success")
        return redirect(url_for("index"))

    flash("Pagamento em processamento. Se você já pagou, aguarde alguns instantes e atualize a página.", "info")
    billing_orders = list_orders_by_user(current_user.id, limit=10)
    billing_history = _history_from_orders(billing_orders)
    return render_template("upgrade_return.html", order=order, billing_history=billing_history)


@analytics_bp.get("/app/upgrade/status")
@require_api_access(require_verified=True)
def upgrade_status():
    token = (request.args.get("token") or "").strip()
    order = get_order_by_token(token)
    if not order:
        return jsonify({"ok": False, "error": "not_found"}), 404
    if not order.user_id or int(order.user_id) != int(current_user.id):
        return jsonify({"ok": False, "error": "forbidden"}), 403

    redirect_to = safe_redirect_path(
        (request.args.get("redirect") or "").strip(),
        allowed_prefixes=("/app/",),
    )
    if order.status == "PAID":
        if apply_paid_order(current_user, order):
            db.session.commit()
        return jsonify(
            {"ok": True, "status": "PAID", "redirect": redirect_to or url_for("index")}
        )

    status = order.status
    if order.billing_id or order.token:
        try:
            remote_status = get_billing_status(order.billing_id, external_id=order.token)
        except AbacatePayError as exc:
            warning = payment_warning_message(str(exc))
            return jsonify({"ok": True, "status": status, "warning": warning})

        if remote_status:
            status = remote_status
            if remote_status == "PAID":
                if order.billing_id:
                    mark_order_paid_by_billing_id(order.billing_id)
                else:
                    mark_order_paid_by_token(order.token)
                order = get_order_by_token(order.token)
                if order and apply_paid_order(current_user, order):
                    db.session.commit()
                return jsonify(
                    {"ok": True, "status": "PAID", "redirect": redirect_to or url_for("index")}
                )

    return jsonify({"ok": True, "status": status})


@analytics_bp.post("/app/subscribe")
@login_required
def subscribe():
    # Endpoint antigo (MVP). Mantido apenas para compatibilidade.
    flash("A troca de plano agora é feita via checkout. Selecione o plano e finalize o pagamento.", "info")
    return redirect(url_for("analytics.upgrade"))


@analytics_bp.get("/app/charts")
@login_required
@require_feature("charts")
def charts_page():
    return render_template("charts.html")


def _safe_int(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _format_month_label(year: int, month: int) -> str:
    if 1 <= month <= 12:
        return f"{MESES[month - 1]} {year}"
    return f"{year}"


def _format_quarter_label(year: int, quarter: int) -> str:
    return f"{quarter}º Trimestre {year}"


def _resolve_charts_period(args) -> dict:
    today = date.today()
    period = (args.get("period") or "month").strip().lower()
    year = _safe_int(args.get("year"), today.year)
    month = _safe_int(args.get("month"), today.month)
    if month < 1 or month > 12:
        month = today.month

    default_quarter = ((month - 1) // 3) + 1
    quarter = _safe_int(args.get("quarter"), default_quarter)
    if quarter < 1 or quarter > 4:
        quarter = default_quarter

    if period == "custom":
        start = _parse_iso_date(args.get("start")) or date(year, month, 1)
        end = _parse_iso_date(args.get("end")) or last_day_of_month(start)
        label = _format_period_label(start, end)
    elif period == "quarter":
        start_month = (quarter - 1) * 3 + 1
        start = date(year, start_month, 1)
        end = last_day_of_month(date(year, start_month + 2, 1))
        label = _format_quarter_label(year, quarter)
    else:
        period = "month"
        start = date(year, month, 1)
        end = last_day_of_month(start)
        label = _format_month_label(year, month)

    if end < start:
        start, end = end, start
        label = _format_period_label(start, end)

    return {
        "type": period,
        "year": year,
        "month": month,
        "quarter": quarter,
        "start": start,
        "end": end,
        "label": label,
    }


def _choose_granularity(period_type: str, start: date, end: date) -> tuple[str, str]:
    if period_type == "quarter":
        return "month", "month"
    if period_type == "custom":
        days = (end - start).days + 1
        if days <= 35:
            return "day", "short"
        return "week", "week"
    return "day", "day"


def _build_buckets(start: date, end: date, granularity: str, label_mode: str) -> list[dict]:
    buckets = []
    if granularity == "month":
        current = date(start.year, start.month, 1)
        while current <= end:
            month_end = last_day_of_month(current)
            bucket_end = month_end if month_end <= end else end
            label = MESES[current.month - 1][:3]
            buckets.append({"start": current, "end": bucket_end, "label": label})
            if current.month == 12:
                current = date(current.year + 1, 1, 1)
            else:
                current = date(current.year, current.month + 1, 1)
        return buckets

    if granularity == "week":
        current = start
        idx = 1
        while current <= end:
            bucket_end = min(current + timedelta(days=6), end)
            buckets.append({"start": current, "end": bucket_end, "label": f"Sem {idx}"})
            idx += 1
            current = bucket_end + timedelta(days=1)
        return buckets

    days = (end - start).days + 1
    for idx in range(days):
        current = start + timedelta(days=idx)
        if label_mode == "short":
            label = current.strftime("%d/%m")
        else:
            label = current.strftime("%d")
        buckets.append({"start": current, "end": current, "label": label})
    return buckets


def _bucket_index_for_date(day: date, start: date, granularity: str) -> int:
    if granularity == "week":
        return (day - start).days // 7
    if granularity == "month":
        return (day.year - start.year) * 12 + (day.month - start.month)
    return (day - start).days


def _summarize_entries(entries: list[Entrada]) -> dict:
    receitas_total = sum(float(e.valor) for e in entries if e.tipo == "receita")
    despesas_total = sum(float(e.valor) for e in entries if e.tipo == "despesa")
    receitas_count = sum(1 for e in entries if e.tipo == "receita")
    despesas_count = sum(1 for e in entries if e.tipo == "despesa")
    saldo_projetado = receitas_total - despesas_total
    economy_pct = round((saldo_projetado / receitas_total) * 100, 1) if receitas_total else 0.0

    return {
        "receitas": round(receitas_total, 2),
        "despesas": round(despesas_total, 2),
        "saldo_projetado": round(saldo_projetado, 2),
        "economy_pct": economy_pct,
        "entradas": len(entries),
        "receitas_count": receitas_count,
        "despesas_count": despesas_count,
    }


def _build_category_breakdown(entries: list[Entrada], entry_type: str) -> list[dict]:
    totals = {}
    total_value = 0.0
    for e in entries:
        if e.tipo != entry_type:
            continue
        cat = _normalize_categoria(getattr(e, "categoria", None))
        totals[cat] = totals.get(cat, 0.0) + float(e.valor)
        total_value += float(e.valor)

    items = []
    for key, total in totals.items():
        if total <= 0:
            continue
        percent = (total / total_value * 100) if total_value else 0.0
        items.append(
            {
                "key": key,
                "label": CATEGORIAS.get(key, "Outros"),
                "total": round(total, 2),
                "percent": round(percent, 1),
            }
        )
    items.sort(key=lambda item: item["total"], reverse=True)
    return items


def _previous_period_meta(period_meta: dict) -> dict:
    period_type = period_meta["type"]
    if period_type == "quarter":
        quarter = period_meta["quarter"]
        year = period_meta["year"]
        if quarter <= 1:
            quarter = 4
            year -= 1
        else:
            quarter -= 1
        start_month = (quarter - 1) * 3 + 1
        start = date(year, start_month, 1)
        end = last_day_of_month(date(year, start_month + 2, 1))
        label = _format_quarter_label(year, quarter)
        return {"start": start, "end": end, "label": label}

    if period_type == "custom":
        start = period_meta["start"]
        end = period_meta["end"]
        days = (end - start).days + 1
        prev_end = start - timedelta(days=1)
        prev_start = prev_end - timedelta(days=days - 1)
        return {"start": prev_start, "end": prev_end, "label": _format_period_label(prev_start, prev_end)}

    year = period_meta["year"]
    month = period_meta["month"]
    if month <= 1:
        year -= 1
        month = 12
    else:
        month -= 1
    start = date(year, month, 1)
    end = last_day_of_month(start)
    return {"start": start, "end": end, "label": _format_month_label(year, month)}


def _delta_payload(current: float, previous: float) -> dict:
    diff = current - previous
    pct = (diff / previous * 100) if previous else None
    return {
        "value": round(diff, 2),
        "pct": round(pct, 1) if pct is not None else None,
    }


def build_period_alerts(
    *,
    user_id: int,
    start: date,
    end: date,
    entries: list[Entrada] | None = None,
    summary: dict | None = None,
    expense_categories: list[dict] | None = None,
) -> list[str]:
    """Monta alertas do periodo (mesma base usada nos graficos/insights)."""

    if entries is None:
        entries = (
            Entrada.query
            .filter(
                Entrada.user_id == user_id,
                Entrada.data >= start,
                Entrada.data <= end,
            )
            .all()
        )

    if summary is None:
        summary = _summarize_entries(entries)

    if expense_categories is None:
        expense_categories = _build_category_breakdown(entries, "despesa")

    alerts: list[str] = []
    if (summary.get("entradas") or 0) <= 0:
        alerts.append("Sem lançamentos no período.")
    else:
        hoje = date.today()
        ref_date = hoje if (start <= hoje <= end) else start
        limite = min(ref_date + timedelta(days=7), end)
        proximas = (
            Entrada.query
            .filter(
                Entrada.user_id == user_id,
                Entrada.tipo == "despesa",
                (Entrada.status.is_(None)) | (Entrada.status != "pago"),
                Entrada.data >= ref_date,
                Entrada.data <= limite,
            )
            .count()
        )
        if proximas:
            alerts.append(f"{proximas} despesas vencem nos próximos 7 dias.")
        else:
            alerts.append("Nenhuma despesa vencendo nos próximos 7 dias.")

    top_expense_category = (expense_categories or [])[:1]
    if top_expense_category:
        top_item = top_expense_category[0]
        if top_item["percent"] >= 35:
            alerts.append(
                f"Categoria {top_item['label']} concentra {top_item['percent']}% das despesas."
            )

    receitas = float(summary.get("receitas") or 0)
    saldo = float(summary.get("saldo_projetado") or 0)
    if saldo < 0:
        alerts.append("Saldo projetado negativo. Ajuste despesas variáveis.")
    elif receitas and saldo < (receitas * 0.1):
        alerts.append("Saldo projetado abaixo de 10% das receitas.")

    return alerts


def _build_charts_insights(summary: dict, categories: dict, highlights: dict, alerts: list[str], statuses: dict) -> dict:
    opportunities = []
    patterns = []

    receitas = float(summary.get("receitas") or 0)
    despesas = float(summary.get("despesas") or 0)
    saldo = float(summary.get("saldo_projetado") or 0)
    economy_pct = summary.get("economy_pct") or 0

    if receitas > 0:
        if economy_pct >= 20:
            opportunities.append(f"Economia acima de {economy_pct:.1f}% das receitas.")
        elif economy_pct <= 5:
            opportunities.append("Economia abaixo de 5%. Ajuste despesas variáveis.")

        if despesas / receitas >= 0.85:
            opportunities.append("Despesas muito próximas das receitas. Reavalie custos fixos.")

    top_expense_category = (categories.get("expense") or [])[:1]
    if top_expense_category:
        top_item = top_expense_category[0]
        patterns.append(f"Maior categoria: {top_item['label']} ({top_item['percent']}% das despesas).")

    pending_total = float(statuses.get("em_andamento") or 0) + float(statuses.get("nao_pago") or 0)
    if pending_total > 0:
        opportunities.append("Existem despesas pendentes. Planeje os próximos pagamentos.")

    best_label = highlights.get("best_bucket_label")
    best_value = highlights.get("best_bucket_total")
    if best_label:
        patterns.append(f"{best_label} foi o melhor período (saldo {best_value:.2f}).")

    return {
        "alerts": alerts[:6],
        "opportunities": opportunities[:6],
        "patterns": patterns[:6],
    }

@analytics_bp.get("/app/charts/data")
@require_api_access(feature="charts")
def charts_data():

    period_meta = _resolve_charts_period(request.args)
    start = period_meta["start"]
    end = period_meta["end"]

    entries = (
        Entrada.query
        .filter(
            Entrada.user_id == current_user.id,
            Entrada.data >= start,
            Entrada.data <= end,
        )
        .all()
    )

    summary = _summarize_entries(entries)

    receitas_antes = (
        db.session.query(func.coalesce(func.sum(Entrada.valor), 0.0))
        .filter(
            Entrada.user_id == current_user.id,
            Entrada.tipo == "receita",
            Entrada.data < start,
        )
        .scalar()
    )
    despesas_pagas_antes = (
        db.session.query(func.coalesce(func.sum(Entrada.valor), 0.0))
        .filter(
            Entrada.user_id == current_user.id,
            Entrada.tipo == "despesa",
            Entrada.status == "pago",
            Entrada.paid_at.isnot(None),
            Entrada.paid_at < start,
        )
        .scalar()
    )
    saldo_anterior = float(receitas_antes) - float(despesas_pagas_antes)
    summary["saldo_anterior"] = round(saldo_anterior, 2)

    granularity, label_mode = _choose_granularity(period_meta["type"], start, end)
    buckets = _build_buckets(start, end, granularity, label_mode)
    bucket_count = len(buckets)

    bucket_receitas = [0.0] * bucket_count
    bucket_despesas = [0.0] * bucket_count

    for e in entries:
        if not e.data:
            continue
        idx = _bucket_index_for_date(e.data, start, granularity)
        if idx < 0 or idx >= bucket_count:
            continue
        if e.tipo == "receita":
            bucket_receitas[idx] += float(e.valor)
        elif e.tipo == "despesa":
            bucket_despesas[idx] += float(e.valor)

    saldo_series = [round(r - d, 2) for r, d in zip(bucket_receitas, bucket_despesas)]
    saldo_acumulado = []
    running = saldo_anterior
    for val in saldo_series:
        running += val
        saldo_acumulado.append(round(running, 2))

    expense_categories = _build_category_breakdown(entries, "despesa")
    income_categories = _build_category_breakdown(entries, "receita")

    status_totais = {"pago": 0.0, "em_andamento": 0.0, "nao_pago": 0.0}
    for e in entries:
        if e.tipo != "despesa":
            continue
        status = (e.status or "em_andamento").strip().lower()
        if status not in STATUS_PADROES:
            status = "em_andamento"
        status_totais[status] += float(e.valor)

    if saldo_series:
        best_idx, best_total = max(enumerate(saldo_series), key=lambda item: item[1])
        best_label = buckets[best_idx]["label"]
    else:
        best_total = 0.0
        best_label = "-"

    top_expense = None
    for e in entries:
        if e.tipo == "despesa":
            if not top_expense or float(e.valor) > float(top_expense.valor):
                top_expense = e
    if top_expense:
        top_expense_total = round(float(top_expense.valor), 2)
        top_expense_label = CATEGORIAS.get(
            _normalize_categoria(getattr(top_expense, "categoria", None)),
            "Outros",
        )
    else:
        top_expense_total = 0.0
        top_expense_label = "-"

    equilibrio = round((summary["despesas"] / summary["receitas"]) * 100, 1) if summary["receitas"] else 0.0

    alerts = build_period_alerts(
        user_id=current_user.id,
        start=start,
        end=end,
        entries=entries,
        summary=summary,
        expense_categories=expense_categories,
    )

    comparison = {"enabled": False}
    if str(request.args.get("compare") or "").lower() in {"1", "true", "yes", "on"}:
        prev_meta = _previous_period_meta(period_meta)
        prev_entries = (
            Entrada.query
            .filter(
                Entrada.user_id == current_user.id,
                Entrada.data >= prev_meta["start"],
                Entrada.data <= prev_meta["end"],
            )
            .all()
        )
        prev_summary = _summarize_entries(prev_entries)
        comparison = {
            "enabled": True,
            "period": {
                "start": prev_meta["start"].isoformat(),
                "end": prev_meta["end"].isoformat(),
                "label": prev_meta["label"],
            },
            "summary": prev_summary,
            "delta": {
                "receitas": _delta_payload(summary["receitas"], prev_summary["receitas"]),
                "despesas": _delta_payload(summary["despesas"], prev_summary["despesas"]),
                "saldo_projetado": _delta_payload(summary["saldo_projetado"], prev_summary["saldo_projetado"]),
                "entradas": _delta_payload(summary["entradas"], prev_summary["entradas"]),
            },
        }

    highlights = {
        "best_bucket_total": round(best_total, 2),
        "best_bucket_label": best_label,
        "top_expense_total": top_expense_total,
        "top_expense_label": top_expense_label,
        "equilibrio": equilibrio,
    }

    insights = _build_charts_insights(
        summary,
        {"expense": expense_categories, "income": income_categories},
        highlights,
        alerts,
        {
            "pago": round(status_totais["pago"], 2),
            "em_andamento": round(status_totais["em_andamento"], 2),
            "nao_pago": round(status_totais["nao_pago"], 2),
        },
    )

    return jsonify(
        {
            "period": {
                "type": period_meta["type"],
                "year": period_meta["year"],
                "month": period_meta["month"],
                "quarter": period_meta["quarter"],
                "start": start.isoformat(),
                "end": end.isoformat(),
                "label": period_meta["label"],
            },
            "summary": summary,
            "comparison": comparison,
            "line": {
                "labels": [bucket["label"] for bucket in buckets],
                "buckets": [
                    {
                        "label": bucket["label"],
                        "start": bucket["start"].isoformat(),
                        "end": bucket["end"].isoformat(),
                    }
                    for bucket in buckets
                ],
                "granularity": granularity,
                "receitas": [round(v, 2) for v in bucket_receitas],
                "despesas": [round(v, 2) for v in bucket_despesas],
                "saldo": saldo_series,
                "saldo_acumulado": saldo_acumulado,
            },
            "categories": {
                "expense": expense_categories,
                "income": income_categories,
            },
            "statuses": {
                "pago": round(status_totais["pago"], 2),
                "em_andamento": round(status_totais["em_andamento"], 2),
                "nao_pago": round(status_totais["nao_pago"], 2),
            },
            "highlights": highlights,
            "insights": insights,
            "updated_at": date.today().isoformat(),
        }
    )


@analytics_bp.get("/app/charts/drilldown")
@require_api_access(feature="charts")
def charts_drilldown():

    start = _parse_iso_date(request.args.get("start"))
    end = _parse_iso_date(request.args.get("end"))
    if not start or not end:
        return json_error("invalid_dates", 422)
    if end < start:
        start, end = end, start

    type_filter = (request.args.get("type") or "all").strip().lower()
    if type_filter not in {"all", "income", "expense"}:
        type_filter = "all"

    category = (request.args.get("category") or "").strip().lower() or None
    limit = _safe_int(request.args.get("limit"), 8)
    if limit < 1:
        limit = 8
    if limit > 30:
        limit = 30

    query = (
        Entrada.query
        .filter(
            Entrada.user_id == current_user.id,
            Entrada.data >= start,
            Entrada.data <= end,
        )
    )

    if type_filter == "income":
        query = query.filter(Entrada.tipo == "receita")
    elif type_filter == "expense":
        query = query.filter(Entrada.tipo == "despesa")

    if category:
        query = query.filter(func.lower(Entrada.categoria) == category)

    entries = (
        query
        .order_by(Entrada.valor.desc(), Entrada.data.desc())
        .limit(limit)
        .all()
    )

    income_total = (
        db.session.query(func.coalesce(func.sum(Entrada.valor), 0.0))
        .filter(
            Entrada.user_id == current_user.id,
            Entrada.tipo == "receita",
            Entrada.data >= start,
            Entrada.data <= end,
        )
        .scalar()
    )
    expense_total = (
        db.session.query(func.coalesce(func.sum(Entrada.valor), 0.0))
        .filter(
            Entrada.user_id == current_user.id,
            Entrada.tipo == "despesa",
            Entrada.data >= start,
            Entrada.data <= end,
        )
        .scalar()
    )

    def _entry_payload(entry: Entrada) -> dict:
        cat = _normalize_categoria(getattr(entry, "categoria", None))
        return {
            "id": entry.id,
            "date": entry.data.isoformat() if entry.data else None,
            "type": "income" if entry.tipo == "receita" else "expense",
            "description": entry.descricao,
            "category": cat,
            "category_label": CATEGORIAS.get(cat, "Outros"),
            "method": _method_label(getattr(entry, "metodo", None)),
            "value": round(float(entry.valor), 2),
            "status": entry.status,
        }

    return jsonify(
        {
            "period": {"start": start.isoformat(), "end": end.isoformat()},
            "filters": {"type": type_filter, "category": category},
            "summary": {
                "income": round(float(income_total or 0.0), 2),
                "expense": round(float(expense_total or 0.0), 2),
                "net": round(float(income_total or 0.0) - float(expense_total or 0.0), 2),
            },
            "items": [_entry_payload(item) for item in entries],
        }
    )


@analytics_bp.get("/app/insights/data")
@analytics_bp.get("/app/compare/data")
@require_api_access(feature="insights")
def insights_data():
    start = _parse_iso_date(request.args.get("start"))
    end = _parse_iso_date(request.args.get("end"))
    if not start or not end:
        return json_error("invalid_dates", 422)
    if end < start:
        return json_error("invalid_range", 422)

    return jsonify(_summary_for_period(start, end))


@analytics_bp.get("/app/insights")
@login_required
@require_feature("insights")
def insights_page():
    return render_template("insights.html")


@analytics_bp.get("/app/compare")
@login_required
def compare_page():
    return redirect(url_for("analytics.insights_page"))


@analytics_bp.get("/app/filters")
@login_required
@require_feature("filters")
def filters_page():
    return render_template("filters.html")


@analytics_bp.get("/app/projection")
@login_required
@require_feature("projection")
def projection_page():
    return render_template("projection.html")



@analytics_bp.route("/app/projection/data", methods=["GET", "POST"])
@require_api_access(feature="projection")
def projection_data():
    payload = request.get_json(silent=True) if request.method == "POST" else request.args
    start_str = (payload.get("start") if payload else None) or None
    end_str = (payload.get("end") if payload else None) or None
    mode = (payload.get("mode") if payload else None) or "cash"

    include_recurring = True
    if payload and payload.get("include_recurring") is not None:
        include_recurring = str(payload.get("include_recurring")).lower() in {"1", "true", "yes", "on"}

    reserve_min = 0.0
    if payload and payload.get("reserve_min") is not None:
        try:
            reserve_min = float(payload.get("reserve_min") or 0.0)
        except (TypeError, ValueError):
            reserve_min = 0.0

    scenario_id = None
    if payload and payload.get("scenario_id"):
        try:
            scenario_id = int(payload.get("scenario_id"))
        except (TypeError, ValueError):
            scenario_id = None

    overrides = payload.get("overrides") if payload else None
    if overrides is None and payload and payload.get("scenario_overrides"):
        overrides = payload.get("scenario_overrides")

    # Se scenario_id vier e overrides não, carregamos do banco
    if scenario_id and overrides is None:
        sc = (
            db.session.query(ProjectionScenario)
            .filter(ProjectionScenario.id == scenario_id, ProjectionScenario.user_id == current_user.id)
            .first()
        )
        if sc:
            try:
                overrides = json.loads(sc.data_json or "{}")
            except Exception:
                overrides = {}

    # datas
    try:
        start_dt = date.fromisoformat(start_str) if start_str else date.today()
    except Exception:
        start_dt = date.today()
    try:
        end_dt = date.fromisoformat(end_str) if end_str else (start_dt + timedelta(days=60))
    except Exception:
        end_dt = start_dt + timedelta(days=60)

    # segurança: não permite período invertido
    if end_dt < start_dt:
        start_dt, end_dt = end_dt, start_dt

    data_out = compute_projection(
        user_id=current_user.id,
        start=start_dt,
        end=end_dt,
        mode=mode,
        include_recurring=include_recurring,
        reserve_min=reserve_min,
        overrides=overrides if isinstance(overrides, dict) else {},
    )

    # devolve também o cenário ativo (para o front persistir)
    data_out["active_overrides"] = overrides if isinstance(overrides, dict) else {}
    return jsonify(data_out)


@analytics_bp.get("/app/projection/scenarios")
@require_api_access(feature="projection")
def projection_scenarios_list():
    rows = (
        db.session.query(ProjectionScenario)
        .filter(ProjectionScenario.user_id == current_user.id)
        .order_by(ProjectionScenario.updated_at.desc())
        .all()
    )
    return jsonify(
        [
            {
                "id": r.id,
                "name": r.name,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None,
            }
            for r in rows
        ]
    )



@analytics_bp.get("/app/projection/scenarios/<int:scenario_id>")
@require_api_access(feature="projection")
def projection_scenarios_get(scenario_id: int):
    sc = (
        db.session.query(ProjectionScenario)
        .filter(ProjectionScenario.id == scenario_id, ProjectionScenario.user_id == current_user.id)
        .first()
    )
    if not sc:
        return jsonify({"error": "Cenário não encontrado"}), 404
    try:
        overrides = json.loads(sc.data_json or "{}")
    except Exception:
        overrides = {}
    return jsonify({"id": sc.id, "name": sc.name, "overrides": overrides, "updated_at": sc.updated_at.isoformat() if sc.updated_at else None})


@analytics_bp.post("/app/projection/scenarios")
@require_api_access(feature="projection")
def projection_scenarios_create():
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()[:80]
    if not name:
        return json_error("nome_obrigatorio", 422)

    overrides = payload.get("overrides") or {}
    try:
        data_json = json.dumps(overrides, ensure_ascii=False)
    except Exception:
        data_json = "{}"

    sc = ProjectionScenario(user_id=current_user.id, name=name, data_json=data_json)
    db.session.add(sc)
    db.session.commit()
    return jsonify({"ok": True, "id": sc.id})


@analytics_bp.put("/app/projection/scenarios/<int:scenario_id>")
@require_api_access(feature="projection")
def projection_scenarios_update(scenario_id: int):
    sc = (
        db.session.query(ProjectionScenario)
        .filter(ProjectionScenario.id == scenario_id, ProjectionScenario.user_id == current_user.id)
        .first()
    )
    if not sc:
        return jsonify({"error": "Cenário não encontrado"}), 404

    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()[:80]
    if name:
        sc.name = name

    overrides = payload.get("overrides")
    if overrides is not None:
        try:
            sc.data_json = json.dumps(overrides, ensure_ascii=False)
        except Exception:
            pass

    db.session.commit()
    return jsonify({"ok": True})


@analytics_bp.delete("/app/projection/scenarios/<int:scenario_id>")
@require_api_access(feature="projection")
def projection_scenarios_delete(scenario_id: int):
    sc = (
        db.session.query(ProjectionScenario)
        .filter(ProjectionScenario.id == scenario_id, ProjectionScenario.user_id == current_user.id)
        .first()
    )
    if not sc:
        return jsonify({"error": "Cenário não encontrado"}), 404

    db.session.delete(sc)
    db.session.commit()
    return jsonify({"ok": True})


@analytics_bp.post("/app/projection/entry/<int:entrada_id>/priority")
@require_api_access(feature="projection")
def projection_entry_priority(entrada_id: int):
    payload = request.get_json(silent=True) or {}
    priority = str(payload.get("priority") or "media").strip().lower()
    if priority not in {"alta", "media", "baixa"}:
        priority = "media"

    e = (
        db.session.query(Entrada)
        .filter(Entrada.id == entrada_id, Entrada.user_id == current_user.id)
        .first()
    )
    if not e:
        return jsonify({"error": "Entrada não encontrada"}), 404

    e.priority = priority
    db.session.commit()
    return jsonify({"ok": True, "priority": priority})


@analytics_bp.get("/app/reports")
@login_required
@require_feature("reports")
def reports_page():
    return render_template("reports.html")


def _build_reports_payload(
    period: str,
    mode: str,
    type_filter: str,
    status_filter: str,
    categories: set[str],
    methods: set[str],
    start_str: str | None,
    end_str: str | None,
    flow_limit: int | None = 500,
    detail: str = "detalhado",
) -> dict:
    start, end = _resolve_report_period(period, start_str, end_str)
    length_days = (end - start).days + 1
    detail = (detail or "detalhado").strip().lower()
    if detail not in {"resumido", "detalhado"}:
        detail = "detalhado"
    is_resumido = detail == "resumido"

    entries = (
        Entrada.query
        .filter(Entrada.user_id == current_user.id)
        .all()
    )

    def matches_filters(entry: Entrada, include_status: bool = True) -> bool:
        if type_filter == "income" and entry.tipo != "receita":
            return False
        if type_filter == "expense" and entry.tipo != "despesa":
            return False

        if categories:
            cat = _normalize_categoria(getattr(entry, "categoria", None))
            if cat not in categories:
                return False

        if not _method_matches(getattr(entry, "metodo", None), methods):
            return False

        if include_status:
            status = (entry.status or "").strip().lower()
            if status_filter == "paid" and status not in {"pago", "recebido"}:
                return False
            if status_filter == "pending" and status in {"pago", "recebido"}:
                return False

        return True

    filtered_all: list[tuple[Entrada, date]] = []
    for entry in entries:
        event_date = _entry_event_date(entry, mode)
        if not event_date:
            continue
        if not matches_filters(entry):
            continue
        filtered_all.append((entry, event_date))

    period_items = [
        (entry, event_date)
        for entry, event_date in filtered_all
        if start <= event_date <= end
    ]

    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=length_days - 1)
    prev_items = [
        (entry, event_date)
        for entry, event_date in filtered_all
        if prev_start <= event_date <= prev_end
    ]

    def totals(items: list[tuple[Entrada, date]]) -> tuple[float, float, float]:
        income = sum(float(e.valor) for e, _ in items if e.tipo == "receita")
        expense = sum(float(e.valor) for e, _ in items if e.tipo == "despesa")
        net = income - expense
        return income, expense, net

    income_total, expense_total, net_total = totals(period_items)
    prev_income, prev_expense, prev_net = totals(prev_items)

    avg_net_values: list[float] = []
    avg_income_values: list[float] = []
    avg_expense_values: list[float] = []
    for i in range(1, 4):
        end_i = start - timedelta(days=length_days * i)
        start_i = end_i - timedelta(days=length_days - 1)
        items_i = [
            (entry, event_date)
            for entry, event_date in filtered_all
            if start_i <= event_date <= end_i
        ]
        income_i, expense_i, net_i = totals(items_i)
        avg_net_values.append(net_i)
        avg_income_values.append(income_i)
        avg_expense_values.append(expense_i)

    avg_net = sum(avg_net_values) / len(avg_net_values) if avg_net_values else 0.0
    avg_income = sum(avg_income_values) / len(avg_income_values) if avg_income_values else 0.0
    avg_expense = sum(avg_expense_values) / len(avg_expense_values) if avg_expense_values else 0.0

    def pct_change(current: float, base: float) -> float | None:
        if not base:
            return None
        return round(((current - base) / abs(base)) * 100, 1)

    economy_pct = round((net_total / income_total) * 100, 1) if income_total else 0.0

    ratio = round((expense_total / income_total) * 100, 1) if income_total else 0.0
    if ratio <= 70:
        health_status = "Equilibrio"
    elif ratio <= 90:
        health_status = "Atencao"
    else:
        health_status = "Critico"

    # Categorias (despesas)
    expense_by_cat: dict[str, float] = {key: 0.0 for key in CATEGORIAS}
    for entry, _ in period_items:
        if entry.tipo != "despesa":
            continue
        cat = _normalize_categoria(getattr(entry, "categoria", None))
        expense_by_cat[cat] = expense_by_cat.get(cat, 0.0) + float(entry.valor)

    prev_expense_by_cat: dict[str, float] = {key: 0.0 for key in CATEGORIAS}
    for entry, _ in prev_items:
        if entry.tipo != "despesa":
            continue
        cat = _normalize_categoria(getattr(entry, "categoria", None))
        prev_expense_by_cat[cat] = prev_expense_by_cat.get(cat, 0.0) + float(entry.valor)

    category_rows = []
    for key, total in expense_by_cat.items():
        if total <= 0:
            continue
        percent = (total / expense_total * 100) if expense_total else 0.0
        prev_total = prev_expense_by_cat.get(key, 0.0)
        delta = pct_change(total, prev_total)
        label = CATEGORIAS.get(key, key.title())
        category_rows.append(
            {
                "key": key,
                "label": label,
                "total": round(total, 2),
                "percent": round(percent, 1),
                "delta": delta,
            }
        )

    category_rows.sort(key=lambda item: item["total"], reverse=True)
    if is_resumido:
        category_rows = category_rows[:8]

    # DRE
    dre_map: dict[str, dict] = {}
    for entry, _ in period_items:
        cat = _normalize_categoria(getattr(entry, "categoria", None))
        label = CATEGORIAS.get(cat, cat.title())
        if cat not in dre_map:
            dre_map[cat] = {"label": label, "income": 0.0, "expense": 0.0}
        if entry.tipo == "receita":
            dre_map[cat]["income"] += float(entry.valor)
        elif entry.tipo == "despesa":
            dre_map[cat]["expense"] += float(entry.valor)

    dre_rows = []
    for _, row in dre_map.items():
        income = float(row["income"])
        expense = float(row["expense"])
        dre_rows.append(
            {
                "label": row["label"],
                "income": round(income, 2),
                "expense": round(expense, 2),
                "net": round(income - expense, 2),
            }
        )
    dre_rows.sort(key=lambda item: abs(item["net"]), reverse=True)

    # Fluxo de caixa
    balance_start = 0.0
    for entry, event_date in filtered_all:
        if event_date < start:
            balance_start += _entry_amount(entry)

    flow_items = sorted(period_items, key=lambda item: (item[1], item[0].id or 0))
    flow_rows = []
    running = balance_start
    if is_resumido:
        current_day = None
        day_income = 0.0
        day_expense = 0.0
        for entry, event_date in flow_items:
            if current_day is None:
                current_day = event_date
            if event_date != current_day:
                flow_rows.append(
                    {
                        "date": current_day.isoformat(),
                        "income": round(day_income, 2),
                        "expense": round(day_expense, 2),
                        "balance": round(running, 2),
                    }
                )
                current_day = event_date
                day_income = 0.0
                day_expense = 0.0

            delta = _entry_amount(entry)
            running += delta
            if entry.tipo == "receita":
                day_income += float(entry.valor)
            elif entry.tipo == "despesa":
                day_expense += float(entry.valor)

        if current_day is not None:
            flow_rows.append(
                {
                    "date": current_day.isoformat(),
                    "income": round(day_income, 2),
                    "expense": round(day_expense, 2),
                    "balance": round(running, 2),
                }
            )

        max_rows = 31 if length_days <= 31 else 45
        if flow_limit:
            max_rows = min(max_rows, flow_limit)
        if len(flow_rows) > max_rows:
            flow_rows = flow_rows[-max_rows:]
    else:
        for entry, event_date in flow_items:
            delta = _entry_amount(entry)
            running += delta
            row = {
                "date": event_date.isoformat(),
                "description": entry.descricao,
                "category": CATEGORIAS.get(_normalize_categoria(entry.categoria), "Outros"),
                "method": _method_label(entry.metodo),
                "status": _status_label(entry.status),
                "income": round(float(entry.valor), 2) if entry.tipo == "receita" else 0.0,
                "expense": round(float(entry.valor), 2) if entry.tipo == "despesa" else 0.0,
                "balance": round(running, 2),
            }
            flow_rows.append(row)

        if flow_limit:
            if len(flow_rows) > flow_limit:
                flow_rows = flow_rows[-flow_limit:]

    # Pendencias
    pending_items = []
    if type_filter != "income" and status_filter != "paid":
        for entry in entries:
            if entry.tipo != "despesa":
                continue
            if (entry.status or "").strip().lower() == "pago":
                continue
            if not matches_filters(entry, include_status=False):
                continue
            if not entry.data:
                continue
            if entry.data < start or entry.data > end:
                continue
            pending_items.append(entry)

    pending_total = sum(float(e.valor) for e in pending_items)
    today = date.today()
    overdue = sum(1 for e in pending_items if e.data and e.data < today)
    due_7 = sum(1 for e in pending_items if e.data and today <= e.data <= today + timedelta(days=7))

    cash_items = _filter_entries_for_period(
        entries,
        start,
        end,
        "cash",
        type_filter,
        status_filter,
        categories,
        methods,
    )
    cash_net = sum(_entry_amount(entry) for entry, _ in cash_items)
    impact_balance = round(cash_net - pending_total, 2)

    pending_rows = []
    for entry in pending_items:
        days_overdue = (today - entry.data).days if entry.data and entry.data < today else 0
        pending_rows.append(
            {
                "date": entry.data.isoformat() if entry.data else None,
                "description": entry.descricao,
                "category": CATEGORIAS.get(_normalize_categoria(entry.categoria), "Outros"),
                "value": round(float(entry.valor), 2),
                "days_overdue": days_overdue,
            }
        )
    pending_rows.sort(key=lambda item: (item.get("value") or 0.0, item.get("days_overdue") or 0), reverse=True)
    if is_resumido:
        pending_rows = pending_rows[:10]

    # Recorrencias (receitas)
    recurrences = (
        Recurrence.query
        .filter(Recurrence.user_id == current_user.id, Recurrence.tipo == "receita")
        .all()
    )
    recurring_items = []
    monthly_estimate = 0.0
    for rec in recurrences:
        exec_count = (
            RecurrenceExecution.query
            .filter(RecurrenceExecution.user_id == current_user.id, RecurrenceExecution.recurrence_id == rec.id)
            .count()
        )
        reliability = min(95, 50 + (exec_count * 5)) if rec.is_enabled else 50
        frequency = rec.frequency or "mensal"
        if frequency == "monthly":
            frequency_label = "Mensal"
        elif frequency == "weekly":
            frequency_label = "Semanal"
        elif frequency == "yearly":
            frequency_label = "Anual"
        else:
            frequency_label = frequency.title()
        frequency_key = (frequency or "").strip().lower()
        if frequency_key in {"weekly", "semanal"}:
            factor = 4.33
        elif frequency_key in {"yearly", "anual"}:
            factor = 1 / 12
        elif frequency_key in {"daily", "diario"}:
            factor = 30
        else:
            factor = 1
        monthly_estimate += float(rec.valor or 0.0) * factor
        recurring_items.append(
            {
                "name": rec.name,
                "frequency": frequency_label,
                "value": round(float(rec.valor or 0.0), 2),
                "reliability": reliability,
            }
        )

    # Alertas
    alerts = []
    if income_total <= 0 and expense_total <= 0:
        alerts.append("Sem lancamentos no periodo.")
    else:
        if ratio >= 95:
            alerts.append("Despesas muito altas em relacao as receitas.")
        elif ratio >= 85:
            alerts.append("Despesas acima da media das receitas.")
        if net_total < 0:
            alerts.append("Resultado liquido negativo no periodo.")
        elif income_total and economy_pct < 10:
            alerts.append("Economia abaixo de 10% das receitas.")
        if category_rows:
            top_cat = category_rows[0]
            if top_cat["percent"] >= 35:
                alerts.append(
                    f"Categoria {top_cat['label']} concentra {top_cat['percent']}% das despesas."
                )
        if pending_total > 0:
            alerts.append(f"{len(pending_items)} pendencias abertas no periodo.")

    comparison_note = ""
    prev_pct = pct_change(net_total, prev_net)
    avg_pct = pct_change(net_total, avg_net)
    income_prev_pct = pct_change(income_total, prev_income)
    expense_prev_pct = pct_change(expense_total, prev_expense)
    income_avg_pct = pct_change(income_total, avg_income)
    expense_avg_pct = pct_change(expense_total, avg_expense)

    if prev_pct is None:
        comparison_note = "Sem base para comparar com o periodo anterior."
    elif prev_pct >= 0:
        comparison_note = "Resultado acima do periodo anterior."
    else:
        comparison_note = "Resultado abaixo do periodo anterior."

    if income_prev_pct is not None and expense_prev_pct is not None:
        if income_prev_pct < 0 and expense_prev_pct > 0:
            comparison_note += " Receitas cairam enquanto despesas subiram."
        elif income_prev_pct > 0 and expense_prev_pct < 0:
            comparison_note += " Receitas subiram e despesas cairam."
        elif expense_prev_pct - income_prev_pct >= 5:
            comparison_note += " Despesas cresceram mais que receitas."

    if avg_pct is not None and abs(avg_pct) >= 5:
        comparison_note += " Comparado a media de 3 periodos, houve variacao relevante."

    return {
        "period": {"start": start.isoformat(), "end": end.isoformat()},
        "summary": {
            "income": round(income_total, 2),
            "expense": round(expense_total, 2),
            "net": round(net_total, 2),
            "economy_pct": economy_pct,
        },
        "comparison": {
            "prev_pct": prev_pct,
            "avg_pct": avg_pct,
            "note": comparison_note,
            "income_prev_pct": income_prev_pct,
            "expense_prev_pct": expense_prev_pct,
            "income_avg_pct": income_avg_pct,
            "expense_avg_pct": expense_avg_pct,
        },
        "health": {
            "ratio": ratio,
            "status": health_status,
            "alerts": alerts[: (8 if detail == "detalhado" else 4)],
        },
        "pending": {
            "count": len(pending_items),
            "total": round(pending_total, 2),
            "impact": impact_balance,
            "overdue": overdue,
            "due_7": due_7,
            "items": pending_rows,
        },
        "dre": {
            "rows": dre_rows,
            "total": {
                "income": round(income_total, 2),
                "expense": round(expense_total, 2),
                "net": round(net_total, 2),
            },
        },
        "flow": {
            "rows": flow_rows,
            "final_balance": round(running, 2),
        },
        "categories": {
            "rows": category_rows,
        },
        "recurring": {
            "items": recurring_items,
            "summary": {
                "count": len(recurring_items),
                "monthly_estimate": round(monthly_estimate, 2),
            },
        },
        "updated_at": date.today().isoformat(),
    }


@analytics_bp.get("/app/reports/data")
@require_api_access(feature="reports")
def reports_data():
    period = (request.args.get("period") or "month").strip().lower()
    mode = (request.args.get("mode") or "cash").strip().lower()
    type_filter = (request.args.get("type") or "all").strip().lower()
    status_filter = (request.args.get("status") or "all").strip().lower()

    categories = _parse_list_param(request.args.get("categories"))
    methods = _parse_list_param(request.args.get("methods"))

    payload = _build_reports_payload(
        period=period,
        mode=mode,
        type_filter=type_filter,
        status_filter=status_filter,
        categories=categories,
        methods=methods,
        start_str=request.args.get("start"),
        end_str=request.args.get("end"),
        flow_limit=500,
        detail="detalhado",
    )
    return jsonify(payload)


@analytics_bp.get("/app/reports/export/pdf")
@require_api_access(feature="reports")
def reports_export_pdf():
    period = (request.args.get("period") or "month").strip().lower()
    mode = (request.args.get("mode") or "cash").strip().lower()
    type_filter = (request.args.get("type") or "all").strip().lower()
    status_filter = (request.args.get("status") or "all").strip().lower()
    detail = (request.args.get("detail") or "resumido").strip().lower()
    if detail not in {"resumido", "detalhado"}:
        detail = "resumido"

    categories = _parse_list_param(request.args.get("categories"))
    methods = _parse_list_param(request.args.get("methods"))
    sections = _parse_list_param(request.args.get("sections"))
    if not sections:
        sections = set(DEFAULT_REPORT_SECTIONS)

    flow_limit = FLOW_LIMIT_RESUMIDO if detail == "resumido" else FLOW_LIMIT_DETALHADO
    payload = _build_reports_payload(
        period=period,
        mode=mode,
        type_filter=type_filter,
        status_filter=status_filter,
        categories=categories,
        methods=methods,
        start_str=request.args.get("start"),
        end_str=request.args.get("end"),
        flow_limit=flow_limit,
        detail=detail,
    )

    period_label = _format_period_label(
        date.fromisoformat(payload["period"]["start"]),
        date.fromisoformat(payload["period"]["end"]),
    )
    mode_label = MODE_LABELS.get(mode, mode.title())
    type_label = "Ambos" if type_filter == "all" else ("Receitas" if type_filter == "income" else "Despesas")
    status_label = "Todos" if status_filter == "all" else ("Pago/Recebido" if status_filter == "paid" else "Pendente")
    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")
    user_name = (getattr(current_user, "full_name", None) or current_user.email)

    logo_path = os.path.join(current_app.root_path, "static", "img", "logo-recorte2.png")
    meta = {
        "title": "Relatorio Financeiro",
        "user_name": user_name,
        "period_label": period_label,
        "mode_label": mode_label,
        "type_label": type_label,
        "status_label": status_label,
        "detail_label": "Resumido" if detail == "resumido" else "Detalhado",
        "generated_at": generated_at,
        "logo_path": logo_path,
    }

    try:
        pdf_bytes = render_reports_pdf(payload, sections, detail, meta)
    except Exception as e:
        current_app.logger.exception("Falha ao gerar PDF de relatorios")
        return jsonify({"error": str(e)}), 500

    download = str(request.args.get("download") or "").lower() in {"1", "true", "yes"}
    buffer = io.BytesIO(pdf_bytes)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=download,
        download_name="relatorio_financeiro.pdf",
    )


@analytics_bp.get("/app/reports/export/excel")
@require_api_access(feature="reports")
def reports_export_excel():
    period = (request.args.get("period") or "month").strip().lower()
    mode = (request.args.get("mode") or "cash").strip().lower()
    type_filter = (request.args.get("type") or "all").strip().lower()
    status_filter = (request.args.get("status") or "all").strip().lower()
    detail = (request.args.get("detail") or "detalhado").strip().lower()
    if detail not in {"resumido", "detalhado"}:
        detail = "detalhado"

    categories = _parse_list_param(request.args.get("categories"))
    methods = _parse_list_param(request.args.get("methods"))
    sections = _parse_list_param(request.args.get("sections"))
    if not sections:
        sections = set(DEFAULT_REPORT_SECTIONS)

    flow_limit = FLOW_LIMIT_RESUMIDO if detail == "resumido" else FLOW_LIMIT_DETALHADO
    payload = _build_reports_payload(
        period=period,
        mode=mode,
        type_filter=type_filter,
        status_filter=status_filter,
        categories=categories,
        methods=methods,
        start_str=request.args.get("start"),
        end_str=request.args.get("end"),
        flow_limit=flow_limit,
        detail=detail,
    )

    def fmt_brl(valor: float) -> str:
        num = float(valor) if valor is not None else 0.0
        return f"R$ {num:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def fmt_date(value: str | None) -> str:
        if not value:
            return ""
        try:
            return date.fromisoformat(value).strftime("%d/%m/%Y")
        except Exception:
            return value

    def sanitize_cell(value):
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            return value
        text = str(value)
        if text and text[0] in {"=", "+", "-", "@"}:
            return "'" + text
        return text

    def sanitize_row(values):
        return [sanitize_cell(v) for v in values]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=";")

        period_label = _format_period_label(
            date.fromisoformat(payload["period"]["start"]),
            date.fromisoformat(payload["period"]["end"]),
        )
        writer.writerow(sanitize_row(["Relatorio financeiro"]))
        writer.writerow(sanitize_row(["Usuario", getattr(current_user, "full_name", None) or current_user.email]))
        writer.writerow(sanitize_row(["Periodo", period_label]))
        writer.writerow(sanitize_row(["Regime", MODE_LABELS.get(mode, mode.title())]))
        writer.writerow([])

        if "summary" in sections:
            writer.writerow(sanitize_row(["Resumo executivo"]))
            writer.writerow(sanitize_row(["Total receitas", fmt_brl(payload["summary"]["income"])]))
            writer.writerow(sanitize_row(["Total despesas", fmt_brl(payload["summary"]["expense"])]))
            writer.writerow(sanitize_row(["Resultado liquido", fmt_brl(payload["summary"]["net"])]))
            writer.writerow(sanitize_row(["% economia", f"{payload['summary']['economy_pct']}%"]))
            writer.writerow([])

        if "dre" in sections:
            writer.writerow(sanitize_row(["DRE"]))
            writer.writerow(sanitize_row(["Categoria", "Receitas", "Despesas", "Resultado"]))
            for row in payload["dre"]["rows"]:
                writer.writerow(sanitize_row([
                    row["label"],
                    fmt_brl(row["income"]),
                    fmt_brl(row["expense"]),
                    fmt_brl(row["net"]),
                ]))
            writer.writerow(sanitize_row([
                "Resultado total",
                fmt_brl(payload["dre"]["total"]["income"]),
                fmt_brl(payload["dre"]["total"]["expense"]),
                fmt_brl(payload["dre"]["total"]["net"]),
            ]))
            writer.writerow([])

        if "flow" in sections:
            writer.writerow(sanitize_row(["Fluxo de caixa"]))
            writer.writerow(sanitize_row(["Data", "Descricao", "Categoria", "Metodo", "Entrada", "Saida", "Saldo"]))
            for row in payload["flow"]["rows"]:
                writer.writerow(sanitize_row([
                    fmt_date(row["date"]),
                    row["description"],
                    row["category"],
                    row["method"],
                    fmt_brl(row["income"]) if row["income"] else "",
                    fmt_brl(row["expense"]) if row["expense"] else "",
                    fmt_brl(row["balance"]),
                ]))
            writer.writerow(sanitize_row(["Saldo final", "", "", "", "", "", fmt_brl(payload["flow"]["final_balance"])]))
            writer.writerow([])

        if "categories" in sections:
            writer.writerow(sanitize_row(["Categorias"]))
            writer.writerow(sanitize_row(["Categoria", "Total", "%", "Variacao"]))
            for row in payload["categories"]["rows"]:
                writer.writerow(sanitize_row([
                    row["label"],
                    fmt_brl(row["total"]),
                    f"{row['percent']}%",
                    f"{row['delta']}%" if row.get("delta") is not None else "",
                ]))
            writer.writerow([])

        if "recurring" in sections:
            writer.writerow(sanitize_row(["Recorrencias (receitas)"]))
            writer.writerow(sanitize_row(["Nome", "Frequencia", "Valor medio", "Confiabilidade"]))
            for item in payload["recurring"]["items"]:
                writer.writerow(sanitize_row([
                    item["name"],
                    item["frequency"],
                    fmt_brl(item["value"]),
                    f"{item['reliability']}%",
                ]))
            writer.writerow([])

        if "pending" in sections:
            writer.writerow(sanitize_row(["Pendencias"]))
            writer.writerow(sanitize_row(["Vencimento", "Descricao", "Categoria", "Valor", "Dias atraso"]))
            for item in payload["pending"]["items"]:
                writer.writerow(sanitize_row([
                    fmt_date(item["date"]),
                    item["description"],
                    item["category"],
                    fmt_brl(item["value"]),
                    item["days_overdue"],
                ]))

        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = "attachment; filename=relatorio_financeiro.csv"
        response.headers["Content-Type"] = "text/csv; charset=utf-8"
        return response

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B7F4A")
    title_font = Font(bold=True, size=13)

    def add_sheet(title: str):
        ws = wb.create_sheet(title)
        ws.page_setup.fitToWidth = 1
        return ws

    def write_row(ws, row_idx: int, values, formats=None, bold=False):
        for col_idx, value in enumerate(values, start=1):
            cell_value = sanitize_cell(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            if formats and col_idx - 1 < len(formats) and formats[col_idx - 1]:
                cell.number_format = formats[col_idx - 1]
            if bold:
                cell.font = Font(bold=True)

    currency_fmt = 'R$ #,##0.00'
    percent_fmt = '0.0%'

    period_label = _format_period_label(
        date.fromisoformat(payload["period"]["start"]),
        date.fromisoformat(payload["period"]["end"]),
    )
    user_label = getattr(current_user, "full_name", None) or current_user.email

    if "summary" in sections:
        ws = add_sheet("Resumo")
        ws["A1"] = "Relatorio financeiro"
        ws["A1"].font = title_font
        ws.append(sanitize_row(["Usuario", user_label]))
        ws.append(sanitize_row(["Periodo", period_label]))
        ws.append(sanitize_row(["Regime", MODE_LABELS.get(mode, mode.title())]))
        ws.append([])
        write_row(ws, 6, ["Resumo executivo"], bold=True)
        ws.append(sanitize_row(["Total receitas", payload["summary"]["income"]]))
        ws.append(sanitize_row(["Total despesas", payload["summary"]["expense"]]))
        ws.append(sanitize_row(["Resultado liquido", payload["summary"]["net"]]))
        ws.append(sanitize_row(["% economia", payload["summary"]["economy_pct"] / 100]))
        ws["B7"].number_format = currency_fmt
        ws["B8"].number_format = currency_fmt
        ws["B9"].number_format = currency_fmt
        ws["B10"].number_format = percent_fmt

    if "dre" in sections:
        ws = add_sheet("DRE")
        headers = ["Categoria", "Receitas", "Despesas", "Resultado"]
        write_row(ws, 1, headers, bold=True)
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
        row_idx = 2
        for row in payload["dre"]["rows"]:
            write_row(ws, row_idx, [row["label"], row["income"], row["expense"], row["net"]])
            for col in (2, 3, 4):
                ws.cell(row=row_idx, column=col).number_format = currency_fmt
            row_idx += 1
        write_row(ws, row_idx, ["Resultado total", payload["dre"]["total"]["income"], payload["dre"]["total"]["expense"], payload["dre"]["total"]["net"]], bold=True)
        for col in (2, 3, 4):
            ws.cell(row=row_idx, column=col).number_format = currency_fmt

    if "flow" in sections:
        ws = add_sheet("Fluxo")
        headers = ["Data", "Descricao", "Categoria", "Metodo", "Entrada", "Saida", "Saldo"]
        write_row(ws, 1, headers, bold=True)
        for col in range(1, 8):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        row_idx = 2
        for row in payload["flow"]["rows"]:
            write_row(ws, row_idx, [
                fmt_date(row["date"]),
                row["description"],
                row["category"],
                row["method"],
                row["income"] or None,
                row["expense"] or None,
                row["balance"],
            ])
            for col in (5, 6, 7):
                ws.cell(row=row_idx, column=col).number_format = currency_fmt
            row_idx += 1
        write_row(ws, row_idx, ["Saldo final", "", "", "", "", "", payload["flow"]["final_balance"]], bold=True)
        ws.cell(row=row_idx, column=7).number_format = currency_fmt

    if "categories" in sections:
        ws = add_sheet("Categorias")
        headers = ["Categoria", "Total", "%", "Variacao"]
        write_row(ws, 1, headers, bold=True)
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        row_idx = 2
        for row in payload["categories"]["rows"]:
            write_row(ws, row_idx, [row["label"], row["total"], row["percent"] / 100, (row["delta"] / 100) if row.get("delta") is not None else None])
            ws.cell(row=row_idx, column=2).number_format = currency_fmt
            ws.cell(row=row_idx, column=3).number_format = percent_fmt
            ws.cell(row=row_idx, column=4).number_format = percent_fmt
            row_idx += 1

    if "recurring" in sections:
        ws = add_sheet("Recorrencias")
        headers = ["Nome", "Frequencia", "Valor medio", "Confiabilidade"]
        write_row(ws, 1, headers, bold=True)
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        row_idx = 2
        for item in payload["recurring"]["items"]:
            write_row(ws, row_idx, [item["name"], item["frequency"], item["value"], item["reliability"] / 100])
            ws.cell(row=row_idx, column=3).number_format = currency_fmt
            ws.cell(row=row_idx, column=4).number_format = percent_fmt
            row_idx += 1

    if "pending" in sections:
        ws = add_sheet("Pendencias")
        headers = ["Vencimento", "Descricao", "Categoria", "Valor", "Dias atraso"]
        write_row(ws, 1, headers, bold=True)
        for col in range(1, 6):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        row_idx = 2
        for item in payload["pending"]["items"]:
            write_row(ws, row_idx, [fmt_date(item["date"]), item["description"], item["category"], item["value"], item["days_overdue"]])
            ws.cell(row=row_idx, column=4).number_format = currency_fmt
            row_idx += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=relatorio_financeiro.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response
